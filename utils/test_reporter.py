import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_FILE = "test_logs.csv"
REPORT_FILE = "产品测试报告.xlsx"

TOKEN_LIMIT = 100000

QWEN_INPUT_PRICE_PER_1K = 0.008
QWEN_OUTPUT_PRICE_PER_1K = 0.02


def calculate_estimated_cost(prompt_tokens, completion_tokens):
    input_cost = prompt_tokens / 1000 * QWEN_INPUT_PRICE_PER_1K
    output_cost = completion_tokens / 1000 * QWEN_OUTPUT_PRICE_PER_1K
    return round(input_cost + output_cost, 4)


def load_logs():
    if not os.path.exists(LOG_FILE):
        return pd.DataFrame()

    df = pd.read_csv(LOG_FILE)

    numeric_cols = [
        "金额",
        "prompt_tokens",
        "completion_tokens",
        "total_tokens",
        "耗时秒"
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def build_round_summary(df):
    if df.empty:
        return pd.DataFrame()

    grouped = df.groupby("测试轮次")

    summary = grouped.agg(
        发票数量=("文件名", "count"),
        成功识别数量=("是否成功", lambda x: (x == "成功").sum()),
        异常数量=("是否成功", lambda x: (x != "成功").sum()),
        合计金额=("金额", "sum"),
        输入Token=("prompt_tokens", "sum"),
        输出Token=("completion_tokens", "sum"),
        总Token=("total_tokens", "sum"),
        平均识别耗时=("耗时秒", "mean")
    ).reset_index()

    summary["成功率"] = summary["成功识别数量"] / summary["发票数量"]
    summary["平均Token/张"] = summary["总Token"] / summary["发票数量"]
    summary["预计成本"] = summary.apply(
        lambda row: calculate_estimated_cost(row["输入Token"], row["输出Token"]),
        axis=1
    )

    return summary[
        [
            "测试轮次",
            "发票数量",
            "成功识别数量",
            "异常数量",
            "成功率",
            "合计金额",
            "总Token",
            "平均Token/张",
            "预计成本",
            "平均识别耗时"
        ]
    ]


def build_total_summary(df):
    if df.empty:
        return pd.DataFrame()

    invoice_count = len(df)
    success_count = len(df[df["是否成功"] == "成功"])
    abnormal_count = invoice_count - success_count

    prompt_tokens = df["prompt_tokens"].sum()
    completion_tokens = df["completion_tokens"].sum()
    total_tokens = df["total_tokens"].sum()

    total = {
        "发票数量": invoice_count,
        "成功识别数量": success_count,
        "异常数量": abnormal_count,
        "成功率": success_count / invoice_count if invoice_count else 0,
        "合计金额": df["金额"].sum(),
        "总Token": total_tokens,
        "平均Token/张": total_tokens / invoice_count if invoice_count else 0,
        "预计成本": calculate_estimated_cost(prompt_tokens, completion_tokens),
        "平均识别耗时": df["耗时秒"].mean(),
        "Token预算上限": TOKEN_LIMIT,
        "Token剩余额度": TOKEN_LIMIT - total_tokens,
        "Token使用率": total_tokens / TOKEN_LIMIT if TOKEN_LIMIT else 0
    }

    return pd.DataFrame([total])


def export_test_report():
    df = load_logs()

    if df.empty:
        return None

    round_summary = build_round_summary(df)
    total_summary = build_total_summary(df)

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        total_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        round_summary.to_excel(writer, sheet_name="分轮测试指标", index=False)
        df.to_excel(writer, sheet_name="测试明细", index=False)

    format_excel(REPORT_FILE)

    return REPORT_FILE


def format_excel(file_path):
    workbook = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column

            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[get_column_letter(column)].width = min(max_length + 4, 40)

    workbook.save(file_path)